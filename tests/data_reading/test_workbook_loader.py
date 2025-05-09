import unittest
import tempfile
import openpyxl
from src.data_reading.workbook_loader import load_workbook

class TestLoadWorkbook(unittest.TestCase):

    def __init__(self, methodName = "runTest"):
        super().__init__(methodName)
        self.tempFile = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "TestValue"
        wb.save(self.tempFile.name)
    
    def test_load_workbook_successfully(self):
        ws = load_workbook(self.tempFile.name)
        self.assertEqual(ws["A1"].value, "TestValue")
    
    def test_invalid_file(self):
        with self.assertRaises(FileNotFoundError):
            load_workbook("non.xlsx")