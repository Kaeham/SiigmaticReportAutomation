import openpyxl.workbook
from src.data_reading.water_reading import extract_water_data, _check_section_pass
import unittest
import openpyxl

class TestWaterExtraction(unittest.TestCase):
    def setUp(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
    
    def test_valid_water_value(self):
        self.ws.cell(row=3, column=23).value = 300  # waterRow
        self.ws.cell(row=5, column=23).value = "Test passed"  # comment row
        self.ws.cell(row=18, column=23).value = "Y"  # sec1a
        self.ws.cell(row=20, column=23).value = "Y"  # sec1b
        self.ws.cell(row=23, column=23).value = "Y"  # sec2[0]
        self.ws.cell(row=24, column=23).value = "Y"  # sec2[1]
        self.ws.cell(row=25, column=23).value = "Y"  # sec2[2]

        result = extract_water_data(self.ws)
        self.assertEqual(result[0], "300")
        self.assertIn("Test passed", result[1])


    def test_invalid_water_value(self):
        self.ws.cell(row=3, column=23).value = "AAAAA"
        result = extract_water_data(self.ws)
        self.assertEqual(result, ("0", "N/A"))
    
    def test_none_water_value(self):
        self.ws.cell(row=3, column=23).value = None  # No value
        result = extract_water_data(self.ws)
        self.assertEqual(result, ("0", "N/A"))

    def test_passed_in_comments(self):
        self.ws.cell(row=3, column=23).value = 250
        self.ws.cell(row=5, column=23).value = "This test passed successfully"
        water, comment = extract_water_data(self.ws)
        self.assertTrue("passed" in comment)

    def test_unformatted_passed_in_comments(self):
        self.ws.cell(row=3, column=23).value = 200
        self.ws.cell(row=5, column=23).value = "PASSED ALL CHECKS"
        result = extract_water_data(self.ws)
        self.assertEqual(result, ("200", "PASSED ALL CHECKS"))

    def test_section_pass(self):
        self.ws.cell(row=3, column=23).value = 400
        self.ws.cell(row=18, column=23).value = "Y"
        self.ws.cell(row=20, column=23).value = "Y"
        self.ws.cell(row=23, column=23).value = "Y"
        self.ws.cell(row=24, column=23).value = "Y"
        self.ws.cell(row=25, column=23).value = "Y"
        result = _check_section_pass(self.ws, 18, 23)
        self.assertTrue(result)


    def test_section_invaid_input(self):
        self.ws.cell(row=3, column=23).value = 400
        self.ws.cell(row=18, column=23).value = 12
        self.ws.cell(row=20, column=23).value = 1
        self.ws.cell(row=23, column=23).value = 1
        self.ws.cell(row=24, column=23).value = 1
        self.ws.cell(row=25, column=23).value = 1
        result = _check_section_pass(self.ws, 18, 23)
        self.assertFalse(result)
    
    def test_section_none_input(self):
        self.ws.cell(row=3, column=23).value = 400
        self.ws.cell(row=18, column=23).value = None
        self.ws.cell(row=20, column=23).value = None
        self.ws.cell(row=23, column=23).value = None
        self.ws.cell(row=24, column=23).value = None
        self.ws.cell(row=25, column=23).value = None
        result = _check_section_pass(self.ws, 18, 23)
        self.assertFalse(result)