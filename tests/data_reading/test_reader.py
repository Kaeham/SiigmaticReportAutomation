import unittest
import os
from openpyxl import Workbook
from tempfile import NamedTemporaryFile
from src.data_reading.reader import load_report_data
from src.data_reading.schema import ReportData

class TestLoadReportData(unittest.TestCase):

    def __init__(self, methodName = "runTest"):
        super().__init__(methodName)
        # Create a temporary Excel workbook with basic structure
        self.wb = Workbook()
        self.ws = self.wb.active

        # Fill in minimal dummy values required by each extraction module
        # Title (rows 2-9, col 2)
        for i in range(8):
            self.ws.cell(row=2+i, column=2, value=f"Value {i+1}")

        self.ws.cell(row=10, column=7, value="sash 1")  # Deflection pos name
        self.ws.cell(row=18, column=11, value=250)      # Deflection pos value
        self.ws.cell(row=38, column=7, value="sash 1")  # Deflection neg name
        self.ws.cell(row=46, column=11, value=245)      # Deflection neg value
        for i in range(1, 4):
            self.ws.cell(row=18, column=11-i).value=2.5
            self.ws.cell(row=46, column=11-i).value=2.5

        self.ws.cell(row=14, column=3, value=800)       # Serviceability

        # Operating force (row 10 + 6, col 14/15)
        self.ws.cell(row=10, column=14, value=12.0)
        self.ws.cell(row=10, column=15, value=5.0)
        self.ws.cell(row=16, column=14, value=10.0)
        self.ws.cell(row=16, column=15, value=3.0)

        # Air infiltration (row 17/18, col 20)
        self.ws.cell(row=17, column=20, value=0.75)
        self.ws.cell(row=18, column=20, value=0.65)

        # Water penetration (row 3, col 23)
        self.ws.cell(row=3, column=23, value=450)
        self.ws.cell(row=5, column=23, value="Passed all conditions")

        # Ultimate (row 3/4, col 25) and observations (rows 6–10 and 15–19)
        self.ws.cell(row=3, column=25, value=5200)
        self.ws.cell(row=4, column=25, value=5000)
        for i in range(5):
            self.ws.cell(row=6+i, column=26, value="Y")
            self.ws.cell(row=15+i, column=26, value="N")

        # Save to a temp file
        self.temp_file = NamedTemporaryFile(delete=False, suffix=".xlsx")
        self.wb.save(self.temp_file.name)

    def test_load_report_data_returns_report_data(self):
        report = load_report_data(self.temp_file.name)
        self.assertIsInstance(report, ReportData)

    def test_report_data_fields_are_correct(self):
        report = load_report_data(self.temp_file.name)

        self.assertEqual(len(report.body_info), 8)
        self.assertEqual(report.deflection_val, '800')
        self.assertEqual(report.filename, os.path.splitext(os.path.basename(self.temp_file.name))[0])

        self.assertEqual(len(report.deflections), 1)
        self.assertEqual(report.deflections[0][1], "250")
        self.assertEqual(report.deflections[0][2], "245")

        self.assertEqual(report.operating_forces[0], ("12.0", "5.0"))
        self.assertEqual(report.air_data, ("0.75", "0.65"))
        self.assertEqual(report.water[0], "450")
        self.assertEqual(len(report.ultimate[2]), 10)

    def test_handles_missing_values_gracefully(self):
        # Clear some cells to simulate partial data
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.cell(row=3, column=25, value=None)  # UST pos pressure
        self.ws.cell(row=3, column=23, value=None)  # Water pressure
        self.wb.save(self.temp_file.name)

        report = load_report_data(self.temp_file.name)

        self.assertIn("N/A", report.ultimate)
        self.assertEqual(report.water[0], "0")



if __name__ == "__main__":
    unittest.main()
