import unittest
from src.data_reading.air_infiltration_reading import *
import openpyxl

class TestAirInfiltrationReading(unittest.TestCase):
    def __init__(self, methodName = "runTest"):
        super().__init__(methodName)
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def test_valid_data(self):
        self.ws.cell(row=17, column=20, value=0.83)   # Positive
        self.ws.cell(row=18, column=20, value=0.72)   # Negative
        result = extract_air_data(self.ws)
        self.assertEqual(result, ("0.83", "0.72"))

    def test_invalid_data(self):
        self.ws.cell(row=17, column=20, value="abc")  # Invalid string
        self.ws.cell(row=18, column=20, value="xyz")
        result = extract_air_data(self.ws)
        self.assertEqual(result, ("N/A", "N/A"))  # May vary based on how you define fallback

    def test_no_negative_data(self):
        self.ws.cell(row=17, column=20, value=0.85)   # Valid positive
        self.ws.cell(row=18, column=20, value=0)      # Zero negative
        result = extract_air_data(self.ws)
        self.assertEqual(result, ("0.85", "N/A"))

    def test_none_data(self):
        # Do not set any value; defaults to None
        result = extract_air_data(self.ws)
        self.assertEqual(result, ("", ""))

if __name__ == "__main__":
    unittest.main()
