from src.data_reading.deflection_reading import *
import unittest
import openpyxl

class TestDeflectionReading(unittest.TestCase):
    def __init__(self, methodName = "runTest"):
        super().__init__(methodName)
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def test_extracting_valid_members(self):
        self.ws.cell(row=10, column=7, value="sash 1")
        self.ws.cell(row=18, column=11, value=230)
        self.ws.cell(row=38, column=7, value="sash 1")
        self.ws.cell(row=46, column=11, value=210)
        for i in range(1, 4):
            self.ws.cell(row=18, column=11-i).value=2.5
            self.ws.cell(row=46, column=11-i).value=2.5

        result = extract_deflection_data(self.ws)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0][1], "230")
        self.assertEqual(result[0][2], "210")

    def test_extracting_invalid_members(self):
        # Arrange: positive but no matching negative
        self.ws.cell(row=10, column=7, value="orphan sash")
        self.ws.cell(row=18, column=11, value=180)
        for i in range(1, 4):
            self.ws.cell(row=18, column=11-i).value=2.5
            self.ws.cell(row=46, column=11-i).value=2.5

        result = extract_deflection_data(self.ws)
        self.assertEqual(result[0][1], "180")
        self.assertEqual(result[0][2], "N/A")

    def test_validate_sensor_valid_data(self):
        # Arrange: all surrounding sensors non-zero
        self.ws.cell(row=18, column=11, value=200)
        for i in range(1, 4):
            self.ws.cell(row=18, column=11-i).value=2.5

        result = validate_sensor(self.ws, 18, 11)
        self.assertEqual(result, "200")

    def test_validate_sensor_invalid_data(self):
        # Arrange: one surrounding sensor = 0
        self.ws.cell(row=18, column=11, value=200)
        self.ws.cell(row=18, column=9, value=0)

        result = validate_sensor(self.ws, 20, 11)
        self.assertEqual(result, "N/A")

    def test_match_member_present(self):
        members = [("a", 1, 1), ("b", 2, 2)]
        idx = match_member_by_name(members, "b")
        self.assertEqual(idx, 1)

    def test_match_member_absent(self):
        members = [("a", 1, 1), ("b", 2, 2)]
        idx = match_member_by_name(members, "c")
        self.assertEqual(idx, -1)

    def test_match_member_invalid_name(self):
        members = [("sash 1", 10, 7)]
        idx = match_member_by_name(members, "")
        self.assertEqual(idx, -1)