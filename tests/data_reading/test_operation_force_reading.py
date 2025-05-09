import unittest
from openpyxl import Workbook
from src.data_reading.operation_force_reading import extract_force_data

class TestOperatingForceReading(unittest.TestCase):
    def setUp(self):
        # Setup runs before each test
        self.wb = Workbook()
        self.ws = self.wb.active

    def test_valid_data(self):
        # Arrange: Put 2 valid rows in worksheet
        self.ws.cell(row=10, column=14, value=12.345)
        self.ws.cell(row=10, column=15, value=6.789)
        self.ws.cell(row=16, column=14, value=9.1)
        self.ws.cell(row=16, column=15, value=4.0)

        result = extract_force_data(self.ws)
        expected = [("12.3", "6.8"), ("9.1", "4.0")]
        self.assertEqual(result, expected)

    def test_invalid_data(self):
        # Arrange: Use values that should be treated as invalid
        self.ws.cell(row=10, column=14, value='#DIV/0!')
        self.ws.cell(row=10, column=15, value=None)
        self.ws.cell(row=16, column=14, value=0)
        self.ws.cell(row=16, column=15, value="")

        result = extract_force_data(self.ws)
        self.assertEqual(result, [(0, 0), (0, 0)])

    def test_none_data(self):
        result = extract_force_data(self.ws)
        self.assertEqual(result, [(0, 0), (0, 0)])
