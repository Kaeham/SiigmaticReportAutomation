import unittest
from openpyxl import Workbook
from src.data_reading.serviceability_reading import get_serviceability_data

class TestServiceability(unittest.TestCase):

    def __init__(self, methodName = "runTest"):
        super().__init__(methodName)
        self.wb = Workbook()
        self.ws = self.wb.active

    def test_valid_input(self):
        # Arrange
        self.ws.cell(row=14, column=3, value=1200)
        result = get_serviceability_data(self.ws)
        self.assertEqual(result, "1200")

    def test_invalid_input(self):
        self.ws.cell(row=14, column=3, value="not_a_number")
        result = get_serviceability_data(self.ws)
        self.assertEqual(result, "N/A")

    def test_none_input(self):
        result = get_serviceability_data(self.ws)
        self.assertEqual(result, "N/A")

if __name__ == "__main__":
    unittest.main()
