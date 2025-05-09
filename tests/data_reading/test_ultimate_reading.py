import unittest
from src.data_reading.ultimate_reading import *
import openpyxl

class TestUltimateReading(unittest.TestCase):
    def __init__(self, methodName = "runTest"):
        super().__init__(methodName)
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
    
    def test_valid_ultimate_data(self):
        # Arrange: set pressures
        self.ws.cell(row=3, column=25, value=5200)
        self.ws.cell(row=4, column=25, value=5000)
        # Add 10 observations
        for i in range(5):
            self.ws.cell(row=6+i, column=26, value="N")
            self.ws.cell(row=15+i, column=26, value="N")

        result = extract_ultimate_data(self.ws)
        self.assertEqual(result[0], "5200")
        self.assertEqual(result[1], "5000")
        self.assertEqual(len(result[2]), 10)
        self.assertEqual(result[2][0], "N")
        self.assertEqual(result[2][5], "N")

    def test_invalid_ultimate_data(self):
        # Arrange: pressures are strings, not valid numbers
        self.ws.cell(row=3, column=25, value="abc")
        self.ws.cell(row=4, column=25, value="xyz")
        # Add dummy observations
        for i in range(10):
            self.ws.cell(row=6+i, column=26, value="")

        result = extract_ultimate_data(self.ws)
        self.assertEqual(result[0], "N/A")
        self.assertEqual(result[1], "N/A")
        self.assertEqual(len(result[2]), 10)
        self.assertEqual(result[2], ["N/A"] * 10)

    def test_no_ultimate_data(self):
        # Arrange: leave pressures unset
        # Act
        result = extract_ultimate_data(self.ws)

        # Assert
        self.assertEqual(result[0], "N/A")
        self.assertEqual(result[1], "N/A")
        self.assertEqual(result[2], ["N/A"] * 10)

if __name__ == "__main__":
    unittest.main()