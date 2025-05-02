import openpyxl
import os
from datetime import datetime


class DataReader():
    def __init__(self, file):
        self.defaultName = os.path.splitext(os.path.basename(file))[0]
        self.workbook = openpyxl.load_workbook(file, data_only=True).active
        self.get_title_data(self.workbook)
        self.get_deflection_data(self.workbook)
        self.get_operational_force(self.workbook)
        self.get_aI_data(self.workbook)
        self.get_water_data(self.workbook)
        self.get_ultimate_data(self.workbook)
        self.final = self.titleData, self.deflectionData, self.oFData, self.AIData, self.waterData, self.ultimateData, self.deflectionVal, self.defaultName
        # print(self.final)

    # Title Data
    def get_title_data(self, workbook):
        """
        Extract the title page data for the test.
        Inputs:
            - workbook: the excel file where the data is stored
        Outputs:
            - titleData: array containing the extracted data.
            - filename: the filename of the excel file
            - tuple containing the titleData and filename
        """
        titleRow = 2
        titleCol = 2

        # extracts: Product Type, Product Name, Test Num, Location, Client
        #           Client, Client Info, Tester, Date Tested
        self.titleData = []
        for idx in range(8):
            self.titleData.append(str(workbook.cell(row=titleRow+idx, column=titleCol).value))
        self.titleData[7] = self.format_date(self.titleData[7])
        
        return (self.titleData, self.defaultName)

    def format_date(self, date):
        """
        Format the date into the proper form of DDMMYYYY
        :Inputs:
            - date: the date extracted from excel, in the form YYYYMMDD HHMMSS
        :Outputs:
            - final: date in the format DDMMYYYY
        """
        print(date)
        if date != None and date.lower() != "None".lower():
            res = datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
            final = res.strftime("%d-%m-%Y")
            return final

        # Serviceablity Design Wind Pressure
    
    # Serviceability Data
    def get_serviceability_data(self, workbook):
        """
        extract the serviceability design wind pressure value
        Inputs:
            - workbook: the excel file where the data is stored
        Outputs:
            - serviceability design wind pressure
        """
        return workbook.cell(row=15, column=3).value

    # Deflection data
    def get_deflection_data(self, workbook):
        """
        get the deflection data from the test
        Inputs:
            - workbook: the excel file where the data is stored
        Outputs:
            - finalData: array containing the positive and negative test pressure result of each structural member
                        alongside the name of the structural member
            - deflectionVal: the deflection value being tested for
            - finalData and deflection val are returned as a tuple.
        """
        posNameStart = (10, 7)
        negNameStart = (38, 7)
        posMembers = self.get_all_members(workbook, posNameStart)
        negMembers = self.get_all_members(workbook, negNameStart)
        self.deflectionVal = workbook.cell(row=14, column=3).value
        self.deflectionData = []
        
        if len(posMembers) != len(negMembers):
            print("Note: Positive pressure test count does not match negative pressure test count")

        for member in posMembers:
            name, row, col = member
            posRes = self.extract_deflection_data(workbook, (row, col))
            self.deflectionData.append((name, posRes, "placeholder"))
        
        for member in negMembers:
            name, row, col = member # get member details
            idx = self.check_name_included(self.deflectionData, name) # get index of the member in the array
            negRes = self.extract_deflection_data(workbook, (row, col))
            curName, posRes, placeholder = self.deflectionData[idx]
            self.deflectionData[idx] = (curName, posRes, negRes)
            print((curName, posRes, negRes))

    def get_all_members(self, workbook, start):
        """
        get all the names used in our list
        Inputs:
            - workbook: the excel file where the data is stored
            - start: the starting coordinate for the loop
        Outputs:
            - arrayOfNames: array containing all the names found.
        """
        # initialize starting name values
        curRow, curCol = start
        arrayOfNames = []
        defaultValues = ["", None, "x sash (01)", "x sash(01)", "X Sash"]
        while curRow <= 1048576:
            currentName = workbook.cell(row=curRow, column=curCol).value
            if type(currentName) == str:
                currentName = currentName.lower()
            default = currentName in defaultValues
            if default:
                break
            idx = self.check_name_included(arrayOfNames, currentName)
            if currentName not in arrayOfNames and not default:
                arrayOfNames.append((currentName, curRow, curCol))
            else:
                arrayOfNames[idx] = (currentName, curRow, curCol)
            curRow += 56
        return arrayOfNames

    def check_name_included(self, array_of_names, value):
        """
        checks whether a given name is included in a array of length 3 tuples
        Inputs:
            - array_of_names: array containing names
            - value: the name to be found in array_of_names
        Outputs:
            - False if value is not found in array_of_names
            - idx: if value is in array_of_names, return the idx where it is found in array_of_names
        """
        length = len(array_of_names)
        for idx in range(length):
            name, row, col = array_of_names[idx]
            if name == value:
                return idx
        return False

    def extract_deflection_data(self, workbook, coords):
        """
        gets the deflection data at the specified coordinates
        Inputs:
            - workbook: the excel file where the data is stored
            - coords: row, cell pair where the data is located
        Outputs:
            - value at coords in the workbook. 
        """
        curRow, curCol = coords
        # print(workbook, (curRow+8, curCol+4))
        return self.confirm_deflection_data(workbook, (curRow+8, curCol+4))

    def confirm_deflection_data(self, workbook, coords):
        """
        Function which checks the given deflection test data for validity
        it performs a basic check on the value of the sensors to confirm validity
        Inputs:
            - workbook: excel sheet with the data
            - coords: tuple of (row, col) of the data
        Outputs:
            - val: the value of the given test
        """
        defRow, col = coords
        val = workbook.cell(row=defRow, column=col).value  # Extract value
        
        # Check if any of the previous three columns contain 0
        for x in range(1, 4):
            if workbook.cell(row=defRow, column=col - (x * 2)).value == 0:
                return "N/A"  # Return "N/A" if invalid data is found
        if type(val) in [int, float]:
            val = "{:.0f}".format(val)
        else:
            return "N/A"
        
        # span = workbook.cell(row=defRow+7, column=col).value
        # spanRatio = workbook.cell(row=defRow+13, column=col-1).value
        # if type(spanRatio) in [int, float]:
        #     spanRatio = "{:.0f}".format(spanRatio)

        # return spanRatio

    # Operation Force (for each specimen)
    def get_operational_force(self, workbook):
        """
        extracts the operational force data
        Inputs:
            - workbook: the excel file where the data is stored
        Outputs:
            - val: tuple containing two tuples.
                    the first tuple contains the initiating and maintaining force for opening
                    the second tuple contains the initiating and maintaining force for closing
        """
        self.oFData = ((0, 0), (0, 0))
        oFData = [(9, 14), (15, 14)]
        oFRow, oFCol = oFData[0]
        oFSecondRow, _ = oFData[1]

        # (("{:.1f}".format(valOne), "{:.1f}".format(valTwo)))
        while True:
            valOne = workbook.cell(row=oFRow, column=oFCol).value
            valTwo = workbook.cell(row=oFRow, column=oFCol+1).value
            valThree = workbook.cell(row=oFSecondRow, column=oFCol).value
            valFour = workbook.cell(row=oFSecondRow, column=oFCol+1).value

            if all(val in ["", None, 0] for val in [valOne, valTwo, valThree, valFour]):
                print()
                return self.oFData
            # print([valOne, valTwo, valThree, valFour])
            if all(isinstance(val, (int, float)) for val in [valOne, valTwo, valThree, valFour]):
                self.oFData = (("{:.1f}".format(valOne), "{:.1f}".format(valTwo)), ("{:.1f}".format(valThree), "{:.1f}".format(valFour)))
            else:
                self.oFData = ((valOne, valTwo), (valThree, valFour))
            oFRow += 15
            oFSecondRow += 15

    # Air infiltration
    def get_aI_data(self, workbook):
        """
        extracts air infiltration data from the excel file
        :Inputs
            - workbook: excel sheet containing all the data
        :Outputs
            - aIData: list of air infiltration data in the 
                    format (positive test value, negative test value)
                    for each test
        """
        # find latest aIValue, record the positive one, and negative if it exists
        start = (17, 20)
        notDefault = True
        aIRow, aICol = start
        self.AIData = ("N/A", "N/A")
        while notDefault:
            posVal = workbook.cell(row=aIRow, column=aICol).value
            negVal = workbook.cell(row=aIRow+1, column=aICol).value
            if (posVal in ["", 0, None]):
                notDefault = False
            else:
                if isinstance(posVal, (int, float)):
                    posVal = "{:.2f}".format(posVal)          
                    if isinstance(negVal, (int, float)):
                        negVal = "{:.2f}".format(negVal)
                    else:
                        negVal = 0
                else:
                    posVal = "N/A"
                    negVal = 0
                self.AIData = (posVal, negVal)
                
            aIRow += 18
        return self.AIData

    # Water Penetration
    def get_water_data(self, workbook):
        """
        Extracts water data
        :Inputs:
            - Workbook: excel workbook containing the data
        :Outputs:
            - extractedData: (water measured value, comments about test, modification details)
        """
        # Initial positions
        positions = [(3, 23), (33, 22), (5, 23)]  # (water value, modifications, comments)
        waterRow, waterCol = positions[0]
        modRow, modCol = positions[1]
        commentsRow, commentsCol = positions[2]
        sectionRow = commentsRow + 13
        self.waterData = False

        # extract first, second and third test values
        # store these values if and only if the test is a valid test
        # store them in val, if they are not default values or empty etc.
        # if the current val is empty, return the previous values
        # otherwise iterate until an empty value is found
        loopIdx = 0
        while waterRow <= 1048576:
            # print(waterRow, modRow, commentsRow)
            waterValue = workbook.cell(row=waterRow, column=waterCol).value
            if loopIdx > 0:
                modVal = workbook.cell(row=modRow, column=modCol).value
            else:
                modVal = ""
            waterComments = ""
            validTest = False
            secOne = False
            secTwo = False
            
            secValues = []
            # sec one
            for idx in range(2):
                secValues.append(workbook.cell(row=(sectionRow+2*idx), column=commentsCol).value)
            # sec two
            # idx = 1 from previous for loop
            secValues.append(workbook.cell(row=(sectionRow+2*(idx+1)+1), column=commentsCol).value)
            for idx in range(3):
                idx += 7
                secValues.append(workbook.cell(row=(sectionRow+idx), column=commentsCol).value)
            
            # check whether test passed each sections
            # sec One
            if secValues[0] == "Y":
                if secValues[1] == "Y":
                    secOne = True
            else:
                secOne = True
            
            # sec Two
            if secValues[2] == "N":
                secTwo = True
            elif secValues[3] == "N":
                secTwo = False
            elif secValues[4] == "N":
                secTwo = False
            elif secValues[5] == "N":
                secTwo = False
            else:
                secTwo = True
            
            if secTwo and secOne:
                validTest = True
                # print(waterRow, modRow, commentsRow)
            
            for rows in range(11):
                val = workbook.cell(row=commentsRow+rows, column=commentsCol).value
                if val not in [None, ""]:
                    waterComments += val + "\n"
                    if "passed" in val.lower():
                        validTest = True
            if waterValue in [None, ""]:
                if self.waterData:
                    return self.waterData
                else:
                    return "N/A"
            elif validTest:
                # print(waterValue, modVal, waterComments, validTest)
                # print(isinstance(waterValue, (int, float)), isinstance(modVal, str), isinstance(waterComments, str))
                if isinstance(waterValue, (int, float)):
                    waterValue = "{:.0f}".format(waterValue)
                    if isinstance(waterComments, str):
                        if isinstance(modVal, str):
                            self.waterData = (waterValue, waterComments, modVal)
                        else:
                            self.waterData = (waterValue, waterComments, "N/A")
                    else:
                        self.waterData = (waterValue, "N/A", "N/A")

            if loopIdx > 0:
                modRow += 41
                waterRow += 41
                commentsRow += 41
            else:
                waterRow += 28
                commentsRow += 41
            loopIdx += 1

    # UST Pressure
    def get_ultimate_data(self, workbook):
        """
        Extract the data for the ultimate strength test from the excel file
        Inputs:
            - workbook: the excel file where the data is stored
        Outputs:
            - ultimateData: list in the form [positive pressure, negative pressure, (positive test observations x5, negative test observations x5)]
        """
        notDefault = True
        fRow, fCol = 3, 25
        obsvRow, obsvCol = 6, 26
        self.ultimateData = ("N/A", "N/A", ["N/A"]*10)

        while notDefault or fRow <= 1048576:
            posTestVal = workbook.cell(row =fRow, column=fCol).value
            negTestVal = workbook.cell(row =fRow+1, column=fCol).value
    
            obsvTable = []
            for idx in range(5):
                obsvVal = workbook.cell(row=obsvRow+idx, column=obsvCol).value
                obsvTable.append(obsvVal if obsvVal != None else "")
            
            obsvRow += idx + 4
            for idx in range(5):
                obsvVal = workbook.cell(row=obsvRow+idx, column=obsvCol).value
                obsvTable.append(obsvVal if obsvVal != None else "")
                
            if posTestVal not in ["", 0, None] and negTestVal not in ["", 0, None]:
                self.ultimateData = ("{:.0f}".format(posTestVal), "{:.0f}".format(negTestVal), obsvTable)
            else:
                return self.ultimateData
            
            # modify rows and columns for the next iteration
            # first iteration, hence slightly different modificaitons
            fRow += 19
            obsvRow += 19
        
        return self.ultimateData

    def get_data(self):
        """
        function to return all extracted values from chosen spread sheet

        Output:
            - [(titledata, filename), (deflectionResults, deflectionVal), operationalForceData,
                airInfiltrationData, waterPenetrationData, ultimateStrengthTestData]
        """
        return self.final

    def get_workbook(self):
        """
        return the workbook that data is being extracted from
        """
        return self.workbook
    
    # Validation with GPT
    def check_with_gpt(self, array_of_values):
        title = array_of_values[0]
        deflection = array_of_values[1]
        oFData = array_of_values[2]
        aIData = array_of_values[3]
        waterData = array_of_values[4]
        ultimateData = array_of_values[5]
        # d = deflection_checks(deflection)
        # o = operating_force_checks(oFData)
        # a = air_infiltration_checks(aIData)
        # w = water_penetration_checks(waterData)
        # u = ultimate_strength_checks(ultimateData)
        # print(d, o, a, w, u)
        # print(u)
